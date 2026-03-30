"""
Manages the Excel workbook that tracks all outreach activity.

Sheet layout (one row per email address tried):
  recruiter_name | recruiter_title | company | role_applied | email_address |
  sent | sent_at | bounced | confirmed_email | subject | notes
"""

import os
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import config

COLUMNS = [
    "recruiter_name",
    "recruiter_title",
    "company",
    "role_applied",
    "email_address",
    "sent",
    "sent_at",
    "bounced",
    "confirmed_email",
    "subject",
    "notes",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_BOUNCE_FILL = PatternFill("solid", fgColor="FFC7CE")
_CONFIRMED_FILL = PatternFill("solid", fgColor="C6EFCE")


def _get_or_create_workbook() -> tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    path = config.excel_file
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Outreach"
        _write_header(ws)

    return wb, ws


def _write_header(ws):
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(16, len(col_name) + 4)


def _col(name: str) -> int:
    return COLUMNS.index(name) + 1


def save_outreach_attempt(
    recruiter_name: str,
    recruiter_title: str,
    company: str,
    role_applied: str,
    email_address: str,
    sent: bool,
    sent_at: Optional[str],
    subject: str,
    notes: str = "",
) -> None:
    """Appends a row for one email address attempt."""
    wb, ws = _get_or_create_workbook()

    row = [None] * len(COLUMNS)
    row[COLUMNS.index("recruiter_name")] = recruiter_name
    row[COLUMNS.index("recruiter_title")] = recruiter_title
    row[COLUMNS.index("company")] = company
    row[COLUMNS.index("role_applied")] = role_applied
    row[COLUMNS.index("email_address")] = email_address
    row[COLUMNS.index("sent")] = "Yes" if sent else "No"
    row[COLUMNS.index("sent_at")] = sent_at or ""
    row[COLUMNS.index("bounced")] = ""
    row[COLUMNS.index("confirmed_email")] = ""
    row[COLUMNS.index("subject")] = subject
    row[COLUMNS.index("notes")] = notes

    ws.append(row)
    wb.save(config.excel_file)


def mark_bounced(email_address: str) -> int:
    """
    Marks all rows with this email_address as bounced.
    Returns number of rows updated.
    """
    wb, ws = _get_or_create_workbook()
    email_col = _col("email_address")
    bounced_col = _col("bounced")
    updated = 0

    for row in ws.iter_rows(min_row=2):
        if row[email_col - 1].value == email_address:
            row[bounced_col - 1].value = "Yes"
            for cell in row:
                cell.fill = _BOUNCE_FILL
            updated += 1

    if updated:
        _infer_confirmed_email(ws)
        wb.save(config.excel_file)

    return updated


def _infer_confirmed_email(ws) -> None:
    """
    For each (recruiter_name, company) group:
    - If exactly ONE email was sent and NOT bounced → mark it as confirmed.
    - If multiple sent and all-but-one bounced → that one is confirmed.
    """
    email_col = _col("email_address")
    sent_col = _col("sent")
    bounced_col = _col("bounced")
    confirmed_col = _col("confirmed_email")
    name_col = _col("recruiter_name")
    company_col = _col("company")

    # Group rows by (name, company)
    groups: dict[tuple, list] = {}
    for row in ws.iter_rows(min_row=2):
        key = (row[name_col - 1].value, row[company_col - 1].value)
        groups.setdefault(key, []).append(row)

    for key, rows in groups.items():
        sent_rows = [r for r in rows if r[sent_col - 1].value == "Yes"]
        not_bounced = [r for r in sent_rows if r[bounced_col - 1].value != "Yes"]

        if len(not_bounced) == 1:
            addr = not_bounced[0][email_col - 1].value
            # Mark all confirmed cells for this group
            for r in rows:
                r[confirmed_col - 1].value = addr
                if r[bounced_col - 1].value != "Yes":
                    for cell in r:
                        cell.fill = _CONFIRMED_FILL


def apply_bounce_results(bounced_addresses: list[str]) -> None:
    """Convenience wrapper to mark a list of bounced addresses."""
    for addr in bounced_addresses:
        count = mark_bounced(addr)
        print(f"  Marked bounced in Excel: {addr} ({count} rows updated)")


def print_confirmed_emails() -> None:
    """Prints a summary of confirmed emails inferred from bounce data."""
    wb, ws = _get_or_create_workbook()
    name_col = _col("recruiter_name")
    company_col = _col("company")
    confirmed_col = _col("confirmed_email")

    seen = set()
    print("\n--- Confirmed emails (inferred from bounce data) ---")
    for row in ws.iter_rows(min_row=2):
        confirmed = row[confirmed_col - 1].value
        if confirmed:
            key = (row[name_col - 1].value, row[company_col - 1].value, confirmed)
            if key not in seen:
                seen.add(key)
                print(f"  {row[name_col-1].value} @ {row[company_col-1].value}  →  {confirmed}")
